"""
dashboard.py  ──  KLUB × Frastea US Leads Dashboard
Zeabur start command: gunicorn main:app (via Procfile)

Changelog:
  v1 (2026-06-12): 初始版本，Flask server-side rendering，含篩選（產業/城市/Tier/有無Email）
  v2 (2026-06-12): 修正 table CSS（width: max-content → 100%）
  v3 (2026-06-12): 新增 CSV 匯出（/export.csv）；篩選器加入「州」；重構 build_where()
  v4 (2026-06-12): 加入「發信狀況」連結（frastracker.zeabur.app）；有 email 的資料優先排序
  v5 (2026-06-30): 補充修改履歷
  v6 (2026-06-30): 新增 /upload-db 端點，可上傳 leads.db 至 persistent volume
  v7 (2026-06-30): 合併雲端資料；以舊有 enriched 資料為基底，加入新增 2057 筆（共 6436 筆）
  v8 (2026-06-30): 部署含聯絡人/職稱/Email/電話/網站/LinkedIn 欄位的完整資料庫
  v9 (2026-06-30): 匯入 Apify Google Places CSV 1070 筆（新增 704、補欄 88），共 7140 筆
  v10 (2026-06-30): header 加入「爬取進度」按鈕，連到 klub-lead-scheduler
  v11 (2026-07-01): 下載改為 Excel xlsx，欄位全中文化，標題藍底白字，自動最適欄寬，凍結首列
"""
import sqlite3
import os
import sys
import csv
import io
import traceback
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from flask import Flask, render_template_string, request, Response, jsonify
except Exception as _e:
    with open("/tmp/startup_error.log", "a") as _f:
        _f.write("=== IMPORT ERROR ===\n")
        traceback.print_exc(file=_f)
    raise

try:
    app = Flask(__name__)
    DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "leads.db"))
except Exception as _e:
    with open("/tmp/startup_error.log", "a") as _f:
        _f.write("=== APP INIT ERROR ===\n")
        traceback.print_exc(file=_f)
    raise

HTML = """
<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<title>Leads Dashboard</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', sans-serif; background: #f0f2f5; color: #333; font-size: 14px; }
header { background: #1a1a2e; color: white; padding: 16px 32px; display:flex; align-items:center; justify-content:space-between; }
header h1 { font-size: 1.4rem; }
header p  { font-size: 0.8rem; color: #aaa; }

.container { max-width: 1400px; margin: 24px auto; padding: 0 20px; }

/* 統計卡片 */
.stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 14px; margin-bottom: 24px; }
.card { background: white; border-radius: 10px; padding: 16px 20px; box-shadow: 0 2px 6px rgba(0,0,0,.07); }
.card .num   { font-size: 2rem; font-weight: 700; color: #1a1a2e; }
.card .label { font-size: 0.78rem; color: #888; margin-top: 4px; }

/* 篩選列 */
.filters { background: white; border-radius: 10px; padding: 16px 20px; margin-bottom: 20px;
           box-shadow: 0 2px 6px rgba(0,0,0,.07); display: flex; flex-wrap: wrap; gap: 12px; align-items: flex-end; }
.filter-group { display: flex; flex-direction: column; gap: 4px; }
.filter-group label { font-size: 0.75rem; color: #666; font-weight: 600; text-transform: uppercase; letter-spacing: .5px; }
.filter-group select,
.filter-group input  { padding: 7px 10px; border: 1px solid #ddd; border-radius: 6px; font-size: 13px;
                        background: white; min-width: 160px; outline: none; }
.filter-group select:focus,
.filter-group input:focus { border-color: #4f8ef7; }
.btn { padding: 7px 18px; border: none; border-radius: 6px; cursor: pointer; font-size: 13px; font-weight: 600; }
.btn-primary { background: #4f8ef7; color: white; }
.btn-primary:hover { background: #3a7ae0; }
.btn-reset   { background: #eee; color: #555; }
.btn-reset:hover { background: #ddd; }

/* 表格 */
.table-wrap { background: white; border-radius: 10px; box-shadow: 0 2px 6px rgba(0,0,0,.07); overflow-x: auto; }
.table-meta { padding: 12px 20px; font-size: 12px; color: #888; border-bottom: 1px solid #f0f0f0; }
table { width: 100%; border-collapse: collapse; table-layout: fixed; }
th { background: #f7f8fa; padding: 8px 10px; text-align: left; font-size: 12px; font-weight: 700;
     color: #555; cursor: pointer; user-select: none; white-space: nowrap; border-bottom: 2px solid #e8e8e8; }
th:hover { background: #eef0f4; }
th .sort-icon { margin-left: 4px; color: #bbb; font-size: 10px; }
th.asc  .sort-icon::after { content: ' ▲'; color: #4f8ef7; }
th.desc .sort-icon::after { content: ' ▼'; color: #4f8ef7; }
th:not(.asc):not(.desc) .sort-icon::after { content: ' ⇅'; }
/* 各欄固定寬度 */
th:nth-child(1), td:nth-child(1)  { width: 160px; } /* 公司 */
th:nth-child(2), td:nth-child(2)  { width: 160px; } /* 產業 */
th:nth-child(3), td:nth-child(3)  { width:  90px; } /* 城市 */
th:nth-child(4), td:nth-child(4)  { width:  50px; } /* 州 */
th:nth-child(5), td:nth-child(5)  { width: 120px; } /* 聯絡人 */
th:nth-child(6), td:nth-child(6)  { width: 150px; } /* 職稱 */
th:nth-child(7), td:nth-child(7)  { width: 190px; } /* Email */
th:nth-child(8), td:nth-child(8)  { width: 120px; } /* 電話 */
th:nth-child(9), td:nth-child(9)  { width:  70px; } /* Tier */
th:nth-child(10),td:nth-child(10) { width:  70px; } /* 評論數 */
th:nth-child(11),td:nth-child(11) { width:  55px; } /* 網站 */
th:nth-child(12),td:nth-child(12) { width:  65px; } /* LinkedIn */
td { padding: 8px 10px; border-bottom: 1px solid #f0f0f0; vertical-align: top;
     white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
tr:hover td { background: #fafbff; }
.badge { display:inline-block; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:600; }
.vip     { background:#fff3cd; color:#856404; }
.general { background:#e8f4fd; color:#0c5499; }
.tag-email { color: #2ecc71; font-size:12px; }
.no-data { text-align: center; padding: 40px; color: #aaa; }
a.link { color: #4f8ef7; text-decoration: none; font-size: 12px; }
a.link:hover { text-decoration: underline; }
</style>
</head>
<body>
<header>
  <div><h1>US Leads Dashboard</h1><p>資料庫：leads.db &nbsp;|&nbsp; {{ now }}</p></div>
  <div style="display:flex;align-items:center;gap:16px;">
    <a href="https://klub-lead-scheduler.zeabur.app/" target="_blank"
       style="background:#1F3864;color:#fff;padding:7px 18px;border-radius:6px;text-decoration:none;font-size:13px;font-weight:700;white-space:nowrap;box-shadow:0 2px 6px rgba(0,0,0,.25);">
      📊 爬取進度
    </a>
    <a href="https://frastracker.zeabur.app/sending" target="_blank"
       style="background:#27ae60;color:#fff;padding:7px 16px;border-radius:6px;text-decoration:none;font-size:13px;font-weight:600;white-space:nowrap;">
      📬 發信狀況
    </a>
    <span style="color:#aaa;font-size:12px;">共 {{ total_all }} 筆資料</span><!-- deploy-v10 -->
  </div>
</header>

<div class="container">
  <!-- 統計卡片 -->
  <div class="stats">
    <div class="card"><div class="num" style="color:#4f8ef7;">{{ total_all }}</div><div class="label">總筆數</div></div>
    <div class="card"><div class="num">{{ with_email }}</div><div class="label">有 Email</div></div>
    <div class="card"><div class="num">{{ with_contact }}</div><div class="label">有聯絡人</div></div>
    <div class="card"><div class="num" style="color:#856404;">{{ vip_count }}</div><div class="label">VIP</div></div>
    <div class="card"><div class="num" style="color:#0c5499;">{{ general_count }}</div><div class="label">GENERAL</div></div>
    <div class="card"><div class="num">{{ city_count }}</div><div class="label">城市數</div></div>
    <div class="card"><div class="num">{{ industry_count }}</div><div class="label">產業數</div></div>
  </div>

  <!-- 篩選列 -->
  <form class="filters" method="GET" action="/">
    <div class="filter-group">
      <label>產業</label>
      <select name="industry">
        <option value="">全部產業</option>
        {% for ind in industries %}
        <option value="{{ ind }}" {% if filters.industry == ind %}selected{% endif %}>{{ ind }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="filter-group">
      <label>城市</label>
      <select name="city">
        <option value="">全部城市</option>
        {% for c in cities %}
        <option value="{{ c }}" {% if filters.city == c %}selected{% endif %}>{{ c }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="filter-group">
      <label>Tier</label>
      <select name="tier">
        <option value="">全部</option>
        <option value="VIP"     {% if filters.tier == 'VIP'     %}selected{% endif %}>VIP</option>
        <option value="GENERAL" {% if filters.tier == 'GENERAL' %}selected{% endif %}>GENERAL</option>
      </select>
    </div>
    <div class="filter-group">
      <label>州</label>
      <select name="state">
        <option value="">全部州</option>
        {% for s in states %}
        <option value="{{ s }}" {% if filters.state == s %}selected{% endif %}>{{ s }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="filter-group">
      <label>Email</label>
      <select name="has_email">
        <option value="">全部</option>
        <option value="1" {% if filters.has_email == '1' %}selected{% endif %}>有 Email</option>
        <option value="0" {% if filters.has_email == '0' %}selected{% endif %}>無 Email</option>
      </select>
    </div>
    <div style="display:flex;gap:8px;align-items:flex-end;">
      <button class="btn btn-primary" type="submit">套用篩選</button>
      <a href="/" class="btn btn-reset">重置</a>
      <button class="btn btn-reset" type="submit" formaction="/export.csv" formtarget="_blank">⬇ 下載 xlsx</button>
    </div>
  </form>

  <!-- 結果表格 -->
  <div class="table-wrap">
    <div class="table-meta">找到 <strong>{{ result_count }}</strong> 筆結果（點擊欄位標題排序）</div>
    <table id="leads-table">
      <thead>
        <tr>
          {% set cols = [('company','公司'),('industry','產業'),('city','城市'),('state','州'),
                         ('contact','聯絡人'),('title','職稱'),('email','Email'),
                         ('phone','電話'),('tier','Tier'),('reviews','評論數')] %}
          {% for col, label in cols %}
          <th data-col="{{ col }}" onclick="sortTable(this)"
              class="{{ 'asc' if sort==col and order=='asc' else 'desc' if sort==col and order=='desc' else '' }}">
            {{ label }}<span class="sort-icon"></span>
          </th>
          {% endfor %}
          <th>網站</th>
          <th>LinkedIn</th>
        </tr>
      </thead>
      <tbody>
        {% if rows %}
          {% for r in rows %}
          <tr>
            <td title="{{ r.company }}">{{ r.company or '' }}</td>
            <td title="{{ r.industry }}">{{ r.industry or '' }}</td>
            <td>{{ r.city or '' }}</td>
            <td>{{ r.state or '' }}</td>
            <td title="{{ r.contact }}">{{ r.contact or '' }}</td>
            <td title="{{ r.title }}">{{ r.title or '' }}</td>
            <td>
              {% if r.email %}
              <span class="tag-email">✉</span> {{ r.email }}
              {% endif %}
            </td>
            <td>{{ r.phone or '' }}</td>
            <td>
              {% if r.tier %}
              <span class="badge {{ 'vip' if r.tier=='VIP' else 'general' }}">{{ r.tier }}</span>
              {% endif %}
            </td>
            <td>{{ r.reviews or '' }}</td>
            <td>{% if r.website %}<a class="link" href="{{ r.website }}" target="_blank">連結</a>{% endif %}</td>
            <td>{% if r.linkedin %}<a class="link" href="{{ r.linkedin }}" target="_blank">連結</a>{% endif %}</td>
          </tr>
          {% endfor %}
        {% else %}
          <tr><td colspan="12" class="no-data">沒有符合條件的資料</td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>

<script>
function sortTable(th) {
  const col   = th.dataset.col;
  const cur   = th.classList.contains('asc') ? 'asc' : th.classList.contains('desc') ? 'desc' : '';
  const next  = cur === 'asc' ? 'desc' : 'asc';
  const url   = new URL(window.location.href);
  url.searchParams.set('sort', col);
  url.searchParams.set('order', next);
  window.location.href = url.toString();
}
</script>
</body>
</html>
"""

ALLOWED_SORT = {"company","industry","city","state","contact","title","email","phone","tier","reviews"}

def query(sql, params=()):
    if not os.path.exists(DB_PATH):
        return []
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        return rows
    except Exception:
        return []

def query_one(sql, params=()):
    if not os.path.exists(DB_PATH):
        return 0
    try:
        conn = sqlite3.connect(DB_PATH)
        val = conn.execute(sql, params).fetchone()
        conn.close()
        return val[0] if val else 0
    except Exception:
        return 0

def build_where(industry, city, state, tier, has_email):
    where, params = [], []
    if industry:
        where.append("industry = ?"); params.append(industry)
    if city:
        where.append("city = ?"); params.append(city)
    if tier:
        where.append("tier = ?"); params.append(tier)
    if state:
        where.append("state = ?"); params.append(state)
    if has_email == "1":
        where.append("email IS NOT NULL AND email != ''")
    elif has_email == "0":
        where.append("(email IS NULL OR email = '')")
    return where, params

@app.route("/")
def index():
    from datetime import datetime

    # 篩選參數
    industry  = request.args.get("industry", "")
    city      = request.args.get("city", "")
    state     = request.args.get("state", "")
    tier      = request.args.get("tier", "")
    has_email = request.args.get("has_email", "")
    sort      = request.args.get("sort", "company")
    order     = request.args.get("order", "asc")
    if sort not in ALLOWED_SORT:
        sort = "company"
    if order not in ("asc", "desc"):
        order = "asc"

    # 統計卡片（全量）
    total_all    = query_one("SELECT COUNT(*) FROM leads")
    with_email   = query_one("SELECT COUNT(*) FROM leads WHERE email IS NOT NULL AND email != ''")
    with_contact = query_one("SELECT COUNT(*) FROM leads WHERE contact IS NOT NULL AND contact != ''")
    vip_count    = query_one("SELECT COUNT(*) FROM leads WHERE tier = 'VIP'")
    general_count= query_one("SELECT COUNT(*) FROM leads WHERE tier = 'GENERAL'")
    city_count   = query_one("SELECT COUNT(DISTINCT city) FROM leads")
    industry_count = query_one("SELECT COUNT(DISTINCT industry) FROM leads")

    # 下拉選單選項
    industries = [r[0] for r in query("SELECT DISTINCT industry FROM leads WHERE industry IS NOT NULL ORDER BY industry")]
    cities     = [r[0] for r in query("SELECT DISTINCT city FROM leads WHERE city IS NOT NULL ORDER BY city")]
    states     = [r[0] for r in query("SELECT DISTINCT state FROM leads WHERE state IS NOT NULL ORDER BY state")]

    # 查詢條件
    where, params = build_where(industry, city, state, tier, has_email)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""SELECT company, industry, city, state, contact, title, email,
                     phone, website, linkedin, tier, reviews
              FROM leads {where_sql}
              ORDER BY (CASE WHEN email IS NOT NULL AND email != '' THEN 0 ELSE 1 END), {sort} {order}
              LIMIT 500"""
    rows = query(sql, params)
    result_count = query_one(f"SELECT COUNT(*) FROM leads {where_sql}", params)

    return render_template_string(HTML,
        total_all=total_all, with_email=with_email, with_contact=with_contact,
        vip_count=vip_count, general_count=general_count,
        city_count=city_count, industry_count=industry_count,
        industries=industries, cities=cities, states=states,
        rows=rows, result_count=result_count,
        filters={"industry": industry, "city": city, "state": state, "tier": tier, "has_email": has_email},
        sort=sort, order=order,
        now=datetime.now().strftime("%Y-%m-%d %H:%M")
    )

@app.route("/export.csv")
def export_csv():
    industry  = request.args.get("industry", "")
    city      = request.args.get("city", "")
    state     = request.args.get("state", "")
    tier      = request.args.get("tier", "")
    has_email = request.args.get("has_email", "")

    where, params = build_where(industry, city, state, tier, has_email)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    sql = f"""SELECT company, industry,
                     (COALESCE(city,'') || CASE WHEN city!='' AND state!='' THEN ', ' ELSE '' END || COALESCE(state,'')) AS address,
                     contact, title, email, phone, website, linkedin, tier, reviews
              FROM leads {where_sql}
              ORDER BY (CASE WHEN email IS NOT NULL AND email != '' THEN 0 ELSE 1 END), company ASC"""
    rows = query(sql, params)

    ZH = ["公司名稱","產業","地址","聯絡人","職稱","電子郵件","電話","網站","LinkedIn","類型","評論數"]
    EN = ["company","industry","address","contact","title","email","phone","website","linkedin","tier","reviews"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    hfill  = PatternFill("solid", fgColor="1F3864")
    hfont  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    halign = Alignment(horizontal="center", vertical="center")
    bfont  = Font(name="Arial", size=10)
    balign = Alignment(vertical="center")

    ws.append(ZH)
    for cell in ws[1]:
        cell.fill = hfill; cell.font = hfont; cell.alignment = halign
    ws.row_dimensions[1].height = 22

    for r in rows:
        ws.append([r.get(k) or "" for k in EN])
        for cell in ws[ws.max_row]:
            cell.font = bfont; cell.alignment = balign

    def _w(v):
        s = str(v or "")
        return sum(2 if ord(c) > 0x2E7F else 1 for c in s)
    for ci, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(_w(c.value) for c in col) + 3, 52)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    from datetime import datetime
    fname = f"leads_export_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return Response(buf.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.route("/debug")
def debug():
    import glob
    db_exists = os.path.exists(DB_PATH)
    count = query_one("SELECT COUNT(*) FROM leads") if db_exists else -1
    files = glob.glob("/app/*.db") + glob.glob("/app/output/*.db")
    return jsonify({
        "DB_PATH": DB_PATH,
        "db_exists": db_exists,
        "count": count,
        "db_files": files,
        "version": "v10"
    })

@app.route("/upload-db", methods=["POST"])
def upload_db():
    PASSWORD = os.environ.get("PASSWORD", "")
    if request.form.get("password") != PASSWORD:
        return jsonify({"error": "unauthorized"}), 403
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    f = request.files["file"]
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    f.save(DB_PATH)
    size = os.path.getsize(DB_PATH)
    return jsonify({"ok": True, "path": DB_PATH, "size": size})

if __name__ == "__main__":
    try:
        port = int(os.environ.get("PORT", 8080))
        print(f"Dashboard 啟動中：http://localhost:{port}")
        app.run(host="0.0.0.0", port=port, debug=False)
    except Exception as _e:
        with open("/tmp/startup_error.log", "a") as _f:
            _f.write("=== RUNTIME ERROR ===\n")
            traceback.print_exc(file=_f)
        raise
