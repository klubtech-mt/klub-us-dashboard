"""
us_lead_enricher_step2_v10.py  ──  KLUB × Frastea
更新：
  1. 新目標客戶條件（全部 VIP）
  2. Google Places 補查 VIP 公司官網
  3. Hunter.io 用官網 domain 補聯絡人
  4. 目標 KeyMan 職稱更新
"""

import os, time, re, requests
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GOOGLE_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "")
APOLLO_API_KEY = os.getenv("APOLLO_API_KEY", "")
HUNTER_API_KEY = os.getenv("HUNTER_API_KEY", "02b906f69a9b94c167d65f3cd68991b0d30833a0")

VIP_EMPLOYEES = 50
OUTPUT_DIR    = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

import glob as _glob
_fixed = sorted(_glob.glob("output/既有名單_整理結果_補查地區_*.xlsx"), reverse=True)
KNOWN_FILE = _fixed[0] if _fixed else "既有名單_整理結果_20260525_0840_1_.xlsx"

# ── 全部都是 VIP 產業 ─────────────────────────
VIP_INDUSTRIES = {
    "food service equipment dealer",
    "bubble tea chain store",
    "coffee cafe chain store",
    "tea house chain store",
    "tea trading company",
    "tea wholesaler",
    "beverage equipment maintenance and service",
    "beverage material store",
    # 相容舊版關鍵字
    "coffee roaster", "industrial roaster",
    "tea beverage chain",
    "equipment dealer", "equipment repair",
    "bubble tea",
}

INDUSTRY_TO_CLASS = {
    "food service equipment dealer":              "供應商-設備",
    "bubble tea chain store":                     "連鎖加盟店",
    "coffee cafe chain store":                    "連鎖加盟店",
    "tea house chain store":                      "連鎖加盟店",
    "tea trading company":                        "供應商-茶葉",
    "tea wholesaler":                             "供應商-茶葉",
    "beverage equipment maintenance and service": "供應商-設備維修",
    "beverage material store":                    "供應商-原物料",
    "coffee roaster":                             "供應商-咖啡",
    "industrial roaster":                         "供應商-咖啡",
    "tea beverage chain":                         "連鎖加盟店",
    "equipment dealer":                           "供應商-設備",
    "equipment repair":                           "供應商-設備維修",
    "bubble tea":                                 "連鎖加盟店",
    "hotel":                                      "飯店",
}

# ── 目標 KeyMan 職稱 ──────────────────────────
TARGET_TITLES = [
    "product manager",
    "purchasing",
    "procurement",
    "business development manager",
    "business development",
    "project manager",
    "buyer",
    "category manager",
    "vendor onboarding manager",
    "vendor manager",
    "owner", "founder", "CEO", "president",
    "director", "general manager", "operations manager",
]

EXCLUDE_KEYWORDS = [
    "personal workshop","franchise","franchisee",
    "cafe","coffee shop","coffee house","coffeehouse",
    "restaurant","bistro","diner","eatery",
    "clark","webstaurant","estella",
]

SENDER_EMAIL = {
    "East":    "mt08@frastea.com",
    "Central": "mt04@frastea.com",
    "West":    "mt04@frastea.com",
}

HDR_FILL     = PatternFill("solid", start_color="1F3864", end_color="1F3864")
VIP_FILL     = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
NEW_VIP_FILL = PatternFill("solid", start_color="FFE0B2", end_color="FFE0B2")
NEW_FILL     = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
HDR_FONT     = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN         = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"),  bottom=Side(style="thin"))

def v(val):
    s = str(val) if val is not None else ""
    return "" if s in ("nan","None","") else s

def _hdr(ws, headers, col_widths=None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.alignment=CENTER; c.border=THIN
    ws.row_dimensions[1].height = 20
    if col_widths:
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

def _row(ws, ri, values, is_vip=False, is_new=False, is_new_vip=False):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=ri, column=col, value=val)
        c.font=BODY_FONT; c.alignment=LEFT; c.border=THIN
        if is_new_vip:   c.fill = NEW_VIP_FILL
        elif is_vip:     c.fill = VIP_FILL
        elif is_new:     c.fill = NEW_FILL

def auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_len:
                    max_len = length
            except: pass
        ws.column_dimensions[col_letter].width = min(50, max(8, max_len + 2))

EAST_STATES = {"NY","NJ","MA","PA","MD","VA","WV","NC","SC","GA","FL",
               "CT","RI","VT","NH","ME","DE","DC"}
WEST_STATES = {"CA","OR","WA","NV","AZ","UT","ID","MT","WY","CO","NM","AK","HI"}

def get_region(state):
    if str(state) in EAST_STATES: return "East"
    if str(state) in WEST_STATES: return "West"
    return "Central"

def get_class(industry):
    return INDUSTRY_TO_CLASS.get(industry.lower(), "其他")

def get_sender_email(region):
    return SENDER_EMAIL.get(region, "mt04@frastea.com")

def assign_tier(industry, employees):
    if industry.lower() in VIP_INDUSTRIES: return "VIP"
    if int(employees or 0) >= VIP_EMPLOYEES: return "VIP"
    return "GENERAL"

def is_excluded(name):
    return any(kw in name.lower() for kw in EXCLUDE_KEYWORDS)

def extract_domain(website):
    if not website: return ""
    return website.replace("https://","").replace("http://","").split("/")[0].strip()

# ══════════════════════════════════════════════
# STEP 1：載入既有名單
# ══════════════════════════════════════════════
def load_existing_all():
    sheets = {}
    if not Path(KNOWN_FILE).exists():
        print(f"  [既有名單] 找不到 {KNOWN_FILE}")
        return sheets
    try:
        xl = pd.read_excel(KNOWN_FILE, sheet_name=None)
        for name, df in xl.items():
            df.columns = [c.strip() for c in df.columns]
            sheets[name] = df
        print(f"  [既有名單] 使用：{KNOWN_FILE}")
        for name, df in sheets.items():
            print(f"    - {name}：{len(df)} 筆")
    except Exception as e:
        print(f"  [既有名單] 讀取失敗：{e}")
    return sheets

# ══════════════════════════════════════════════
# STEP 2：排除 + 去重
# ══════════════════════════════════════════════
def dedup_and_exclude(new_df, sheets):
    excluded_rows, keep_mask = [], []
    for _, row in new_df.iterrows():
        name = str(row.get("公司","")).strip()
        if is_excluded(name):
            excluded_rows.append({
                "公司": name,
                "官網/Email": v(row.get("官網","")) or v(row.get("Email","")),
                "來源分頁": "Google Places",
                "排除原因": "符合排除關鍵字",
            })
            keep_mask.append(False)
        else:
            keep_mask.append(True)

    new_df = new_df[keep_mask].reset_index(drop=True)
    print(f"  [排除] 排除 {len(excluded_rows)} 筆，剩餘 {len(new_df)} 筆")

    existing_df = sheets.get("聯絡人明細", pd.DataFrame())
    if not existing_df.empty:
        existing_cos = set(str(r).strip().lower()
                          for r in existing_df.get("公司", pd.Series()).dropna())
        before = len(new_df)
        new_df = new_df[~new_df["公司"].str.strip().str.lower().isin(existing_cos)]
        new_df = new_df.reset_index(drop=True)
        print(f"  [去重] 移除 {before-len(new_df)} 筆，剩餘 {len(new_df)} 筆")

    return new_df, excluded_rows

# ══════════════════════════════════════════════
# STEP 3：Google Places 補查官網
# ══════════════════════════════════════════════
def google_get_website(brand: str, city: str = "", state: str = "") -> str:
    if not GOOGLE_API_KEY: return ""
    # 先用 Text Search 找 place_id
    url    = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    query  = f"{brand} {city} {state}".strip() if city else brand
    params = {"query": query, "language": "en", "key": GOOGLE_API_KEY}
    try:
        resp    = requests.get(url, params=params, timeout=10)
        results = resp.json().get("results", [])
        if not results: return ""
        place_id = results[0].get("place_id","")
        if not place_id: return ""

        # 用 Place Details 取官網
        detail_url = "https://maps.googleapis.com/maps/api/place/details/json"
        detail_params = {
            "place_id": place_id,
            "fields":   "website,formatted_phone_number",
            "key":      GOOGLE_API_KEY,
        }
        detail_resp = requests.get(detail_url, params=detail_params, timeout=10)
        result      = detail_resp.json().get("result", {})
        return result.get("website", "")
    except Exception as e:
        print(f"    [Google Places Detail] {brand}: {e}")
        return ""

# ══════════════════════════════════════════════
# STEP 4：Apollo 查員工數
# ══════════════════════════════════════════════
def apollo_org(company_name, domain=""):
    if not APOLLO_API_KEY: return {}
    url     = "https://api.apollo.io/v1/organizations/search"
    headers = {"Content-Type": "application/json", "X-Api-Key": APOLLO_API_KEY}
    payload = {"q_organization_name": company_name, "page": 1, "per_page": 1}
    if domain: payload["q_organization_domains"] = [domain]
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=15)
        if resp.status_code != 200: return {}
        orgs = resp.json().get("organizations", [])
        if not orgs: return {}
        org = orgs[0]
        return {
            "employees":    org.get("estimated_num_employees") or 0,
            "org_website":  org.get("website_url", ""),
            "org_linkedin": org.get("linkedin_url", ""),
        }
    except: return {}

# ══════════════════════════════════════════════
# STEP 4b：Apollo 查聯絡人電話
# ══════════════════════════════════════════════
def apollo_people_search(company_name: str, domain: str = "") -> list[dict]:
    if not APOLLO_API_KEY: return []
    url = "https://api.apollo.io/v1/mixed_people/api_search"
    headers = {"Content-Type": "application/json", "X-Api-Key": APOLLO_API_KEY}
    payload = {
        "q_organization_name": company_name,
        "person_titles": TARGET_TITLES,
        "page": 1,
        "per_page": 3,
    }
    if domain:
        payload["q_organization_domains"] = [domain]
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=15)
        if resp.status_code != 200: return []
        people = resp.json().get("people", [])
        results = []
        for p in people:
            phone = ""
            # 取第一個有效電話
            for ph in (p.get("phone_numbers") or []):
                if ph.get("sanitized_number"):
                    phone = ph["sanitized_number"]
                    break
            results.append({
                "聯絡人": f"{p.get('first_name','')} {p.get('last_name','')}".strip(),
                "職稱":   p.get("title", ""),
                "Email":  p.get("email", "") or "",
                "電話":   phone,
                "LinkedIn": p.get("linkedin_url", "") or "",
            })
        return results
    except Exception as e:
        print(f"    [Apollo People] {company_name}: {e}")
        return []

# ══════════════════════════════════════════════
# STEP 5：Hunter.io 補聯絡人
# ══════════════════════════════════════════════
def hunter_domain_search(domain: str) -> list[dict]:
    if not HUNTER_API_KEY or not domain: return []
    url    = "https://api.hunter.io/v2/domain-search"
    params = {
        "domain":  domain,
        "api_key": HUNTER_API_KEY,
        "limit":   10,
        "type":    "personal",
    }
    try:
        resp = requests.get(url, params=params, timeout=15)
        if resp.status_code != 200: return []
        emails = resp.json().get("data", {}).get("emails", [])

        # 過濾目標職稱
        contacts = []
        for e in emails:
            position = e.get("position","").lower()
            # 優先目標職稱
            is_target = any(t in position for t in TARGET_TITLES)
            contacts.append({
                "聯絡人":   f"{e.get('first_name','')} {e.get('last_name','')}".strip(),
                "職稱":     e.get("position",""),
                "Email":    e.get("value",""),
                "LinkedIn": e.get("linkedin","") or "",
                "confidence": e.get("confidence", 0),
                "is_target":  is_target,
            })

        # 目標職稱優先，其次信心度
        contacts.sort(key=lambda x: (0 if x["is_target"] else 1, -x["confidence"]))
        return contacts[:3]
    except Exception as e:
        print(f"    [Hunter] {domain}: {e}")
        return []

# ══════════════════════════════════════════════
# STEP 6：主要補充流程
# ══════════════════════════════════════════════
def enrich_all(df: pd.DataFrame) -> list[dict]:
    print(f"\n[Step 3-5] 補查官網 + Apollo員工數 + Hunter補聯絡人（共 {len(df)} 筆）...")
    detail_rows, enriched = [], 0

    for i, row in enumerate(df.to_dict("records"), 1):
        brand    = v(row.get("公司",""))
        industry = v(row.get("產業",""))
        website  = v(row.get("官網",""))
        city     = v(row.get("城市",""))
        state    = v(row.get("State",""))
        region   = v(row.get("東西岸",""))

        # Apollo 查員工數 + 官網
        domain   = extract_domain(website)
        org_info = apollo_org(brand, domain)
        employees     = org_info.get("employees", 0)
        final_website = org_info.get("org_website","") or website
        final_linkedin= org_info.get("org_linkedin","")

        # VIP 判定
        tier = assign_tier(industry, employees)

        # VIP 才補查官網和聯絡人
        contacts = []
        if tier == "VIP":
            # 若無官網，用 Google Places 補查
            if not final_website:
                final_website = google_get_website(brand, city, state)
                time.sleep(0.2)

            domain = extract_domain(final_website)
            if domain:
                contacts = hunter_domain_search(domain)
                time.sleep(0.3)

            # Apollo People 查聯絡人電話
            apollo_people = apollo_people_search(brand, domain)
            # 補電話到 contacts，或新增聯絡人
            apollo_map = {p["聯絡人"].lower(): p for p in apollo_people}
            for c in contacts:
                ap = apollo_map.get(c["聯絡人"].lower(), {})
                if ap.get("電話"):
                    c["電話"] = ap["電話"]
                if not c.get("Email") and ap.get("Email"):
                    c["Email"] = ap["Email"]
            # 若 Hunter 沒有找到，用 Apollo People
            if not contacts and apollo_people:
                contacts = [{
                    "聯絡人": p["聯絡人"], "職稱": p["職稱"],
                    "Email": p["Email"], "電話": p["電話"],
                    "LinkedIn": p["LinkedIn"] or final_linkedin,
                    "confidence": 0, "is_target": True,
                    "來源": "Apollo",
                } for p in apollo_people]
            time.sleep(0.2)

        base = {
            "公司": brand, "產業": industry,
            "分類": get_class(industry),
            "城市": city, "State": state, "東西岸": region,
            "電話": v(row.get("電話","")),
            "官網": final_website,
            "寄件信箱": get_sender_email(region),
            "Tier": tier, "employees": employees,
            "連鎖品牌": "", "建議總部HQ": "",
            "HQ網站": "", "HQ聯絡": "",
        }

        if contacts:
            enriched += 1
            for c in contacts:
                src = c.get("來源") or f"Hunter.io({c.get('confidence',0)}%)"
                detail_rows.append({**base,
                    "聯絡人": c["聯絡人"], "職稱": c["職稱"],
                    "Email":  c["Email"],
                    "電話":   c.get("電話","") or base["電話"],
                    "LinkedIn": c.get("LinkedIn","") or final_linkedin,
                    "來源分頁": src,
                })
        else:
            detail_rows.append({**base,
                "聯絡人": "", "職稱": "",
                "Email":  v(row.get("Email","")),
                "LinkedIn": final_linkedin,
                "來源分頁": "Google Places",
            })

        if i % 50 == 0:
            print(f"  進度：{i}/{len(df)}  VIP有聯絡人：{enriched} 家")
        time.sleep(0.2)

    vip = sum(1 for r in detail_rows if r["Tier"]=="VIP")
    print(f"  完成！VIP：{vip}  GENERAL：{len(detail_rows)-vip}  有聯絡人：{enriched} 家")
    return detail_rows

# ══════════════════════════════════════════════
# STEP 7：輸出 xlsx
# ══════════════════════════════════════════════
def get_next_version():
    existing = list(OUTPUT_DIR.glob("us_leads_最新爬取資料彙總_*.xlsx"))
    if not existing: return 1
    versions = [int(m.group(1)) for f in existing
                if (m := re.search(r"_v(\d+)\.xlsx$", f.name))]
    return max(versions, default=0) + 1

def sort_rows(rows):
    ro = {"East":0,"Central":1,"West":2}
    to = {"VIP":0,"GENERAL":1}
    return sorted(rows, key=lambda r: (
        to.get(r.get("Tier",""),1),
        ro.get(r.get("東西岸",""),9),
        r.get("公司",""),
    ))

def export_xlsx(sheets, new_detail, excluded, ts, version):
    exist_detail   = sheets.get("聯絡人明細",       pd.DataFrame())
    exist_summary  = sheets.get("公司彙整",         pd.DataFrame())
    exist_chain    = sheets.get("連鎖加盟對應總部", pd.DataFrame())
    exist_hunter   = sheets.get("具名窗口(Hunter)", pd.DataFrame())
    exist_excluded = sheets.get("已排除清單",       pd.DataFrame())

    new_detail = sort_rows(new_detail)

    company_map = {}
    for r in new_detail:
        co = r["公司"]
        if co not in company_map:
            company_map[co] = {"count":0,"contacts":0,"emails":[],"r":r}
        company_map[co]["count"] += 1
        if r.get("聯絡人"): company_map[co]["contacts"] += 1
        if r.get("Email"):  company_map[co]["emails"].append(r["Email"])

    new_summary = sort_rows([{
        "Tier":        info["r"].get("Tier","GENERAL"),
        "公司(正規化)": co,
        "分類":        info["r"].get("分類",""),
        "產業":        info["r"].get("產業",""),
        "名單筆數":    info["count"],
        "美國據點數":  "",
        "聯絡人數":    info["contacts"],
        "State":       info["r"].get("State",""),
        "東西岸":      info["r"].get("東西岸",""),
        "寄件信箱":    info["r"].get("寄件信箱",""),
        "連鎖品牌":    "",
        "建議總部HQ":  "",
        "HQ網站":      "",
        "HQ聯絡":      "",
        "官網":        info["r"].get("官網",""),
        "代表Email":   info["emails"][0] if info["emails"] else "",
    } for co, info in company_map.items()])

    new_excluded_rows = [{
        "公司": r.get("公司",""),
        "官網/Email": r.get("官網/Email",""),
        "來源分頁": r.get("來源分頁",""),
        "排除原因": r.get("排除原因",""),
    } for r in excluded]

    wb = Workbook()

    # Sheet 1：公司彙整
    ws1 = wb.active; ws1.title = "公司彙整"
    sum_h = ["Tier","公司(正規化)","分類","產業","名單筆數","美國據點數",
             "聯絡人數","State","東西岸","寄件信箱","連鎖品牌",
             "建議總部HQ","HQ網站","HQ聯絡","官網","代表Email"]
    _hdr(ws1, sum_h)
    ri = 2
    for _, row in exist_summary.iterrows():
        _row(ws1, ri, [v(row.get(h,"")) for h in sum_h],
             is_vip=(str(row.get("Tier","")).upper()=="VIP"))
        ri += 1
    for r in new_summary:
        is_new_vip = r.get("Tier")=="VIP"
        _row(ws1, ri, [r.get(h,"") for h in sum_h],
             is_new_vip=is_new_vip, is_new=(not is_new_vip))
        ri += 1

    # Sheet 2：聯絡人明細
    ws2 = wb.create_sheet("聯絡人明細")
    det_h = ["Tier","公司","分類","產業","聯絡人","職稱",
             "Email","電話","State","東西岸","寄件信箱","官網","LinkedIn","來源分頁"]
    _hdr(ws2, det_h)
    ri = 2
    for _, row in exist_detail.iterrows():
        _row(ws2, ri, [v(row.get(h,"")) for h in det_h],
             is_vip=(str(row.get("Tier","")).upper()=="VIP"))
        ri += 1
    for r in new_detail:
        is_new_vip = r.get("Tier")=="VIP"
        _row(ws2, ri, [r.get(h,"") for h in det_h],
             is_new_vip=is_new_vip, is_new=(not is_new_vip))
        ri += 1

    # Sheet 3：連鎖加盟對應總部
    ws3 = wb.create_sheet("連鎖加盟對應總部")
    chain_h = ["加盟店(原名)","State","連鎖品牌","美國據點數","VIP判定",
               "總部具名窗口(Hunter)","建議總部HQ","HQ網站","HQ聯絡","據點數來源"]
    _hdr(ws3, chain_h)
    ri = 2
    for _, row in exist_chain.iterrows():
        _row(ws3, ri, [v(row.get(h,"")) for h in chain_h])
        ri += 1

    # Sheet 4：已排除清單
    ws4 = wb.create_sheet("已排除清單")
    excl_h = ["公司","官網/Email","來源分頁","排除原因"]
    _hdr(ws4, excl_h)
    ri = 2
    for _, row in exist_excluded.iterrows():
        _row(ws4, ri, [v(row.get(h,"")) for h in excl_h])
        ri += 1
    for r in new_excluded_rows:
        _row(ws4, ri, [r.get(h,"") for h in excl_h])
        ri += 1

    # Sheet 5：具名窗口(Hunter)
    ws5 = wb.create_sheet("具名窗口(Hunter)")
    hunter_h = ["公司/總部","連鎖品牌","Tier","聯絡人","職稱","Email","可信度%","來源"]
    _hdr(ws5, hunter_h)
    ri = 2
    for _, row in exist_hunter.iterrows():
        _row(ws5, ri, [v(row.get(h,"")) for h in hunter_h])
        ri += 1
    for r in new_detail:
        if r.get("Email") and r.get("聯絡人"):
            conf = ""
            m = re.search(r'\((\d+)%\)', r.get("來源分頁",""))
            if m: conf = m.group(1)
            _row(ws5, ri, [
                r.get("公司",""), r.get("連鎖品牌",""), r.get("Tier",""),
                r.get("聯絡人",""), r.get("職稱",""), r.get("Email",""),
                conf, "Hunter.io",
            ], is_new_vip=(r.get("Tier")=="VIP"))
            ri += 1

    # Sheet 6：摘要
    ws6 = wb.create_sheet("摘要")
    ws6["A1"] = "KLUB × Frastea 既有名單 整理結果（驗證用）"
    ws6["A1"].font = Font(name="Arial", bold=True, size=13)
    ws6["A2"] = f"產生時間 {ts}　｜　VIP：所有目標產業皆為VIP 或 員工數≥{VIP_EMPLOYEES}"
    ws6["A3"] = "VIP 產業：Food Service Equipment / Bubble Tea / Coffee Cafe / Tea House / Tea Trading / Tea Wholesaler / Bev Equipment Maintenance / Bev Material"
    ws6["A5"]  = "合併後公司彙整";    ws6["B5"]  = len(exist_summary)+len(new_summary)
    ws6["A6"]  = "  既有名單";        ws6["B6"]  = len(exist_summary)
    ws6["A7"]  = "  本次新增";        ws6["B7"]  = len(new_summary)
    ws6["A8"]  = "本次新增 VIP";      ws6["B8"]  = sum(1 for r in new_summary if r.get("Tier")=="VIP")
    ws6["A9"]  = "本次新增 GENERAL";  ws6["B9"]  = sum(1 for r in new_summary if r.get("Tier")=="GENERAL")
    ws6["A10"] = "本次新增聯絡人";    ws6["B10"] = len(new_detail)
    ws6["A11"] = "本次排除筆數";      ws6["B11"] = len(new_excluded_rows)
    ws6["A13"] = "顏色說明"
    ws6["B13"] = "黃色=既有VIP　橘色=本次新增VIP　淺綠色=本次新增GENERAL"
    for cell in ["A5","A6","A7","A8","A9","A10","A11","A13"]:
        ws6[cell].font = Font(name="Arial", bold=True, size=10)
    ws6.column_dimensions["A"].width = 22
    ws6.column_dimensions["B"].width = 60

    for ws in wb.worksheets:
        auto_col_width(ws)

    fname = OUTPUT_DIR / f"us_leads_最新爬取資料彙總_{ts}_v{version}.xlsx"
    wb.save(str(fname))
    print(f"\n✅ 輸出：{fname}")
    print(f"   公司彙整：{len(exist_summary)+len(new_summary)} 筆（既有{len(exist_summary)}+新增{len(new_summary)}）")
    print(f"   聯絡人明細：{len(exist_detail)+len(new_detail)} 筆")
    print(f"   新增 VIP：{sum(1 for r in new_summary if r.get('Tier')=='VIP')}  GENERAL：{sum(1 for r in new_summary if r.get('Tier')=='GENERAL')}")
    return fname

# ══════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════
def run(csv_path=None):
    import glob as _g
    if not csv_path:
        files = sorted(_g.glob("us_brands_*.csv"), reverse=True)
        if not files:
            print("❌ 找不到 us_brands_*.csv"); return
        csv_path = files[0]
    print(f"[Step2 v9] 讀取：{csv_path}")

    new_df = pd.read_csv(csv_path)
    new_df.columns = [c.strip() for c in new_df.columns]

    rename_map = {
        "Brand Name":"公司","Industry":"產業","City":"城市",
        "Phone":"電話","Website":"官網",
    }
    for eng, chn in rename_map.items():
        if eng in new_df.columns and chn not in new_df.columns:
            new_df = new_df.rename(columns={eng: chn})

    if "東西岸" not in new_df.columns:
        new_df["東西岸"] = new_df["State"].apply(get_region) if "State" in new_df.columns else ""

    sheets           = load_existing_all()
    new_df, excluded = dedup_and_exclude(new_df, sheets)
    detail_rows      = enrich_all(new_df)

    ts      = datetime.now().strftime("%Y-%m-%d %H:%M")
    ts_file = datetime.now().strftime("%Y%m%d_%H%M")
    version = get_next_version()
    export_xlsx(sheets, detail_rows, excluded, ts_file, version)

if __name__ == "__main__":
    import sys
    csv_path = sys.argv[1] if len(sys.argv) > 1 else None
    run(csv_path)
